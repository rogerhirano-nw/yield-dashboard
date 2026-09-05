#!/usr/bin/env python3
"""
Build docs/Newsweek_Digital_Ads_Benchmarks_Master.xlsx — the performance-benchmark
companion to the sales team's Newsweek_Digital_Ads_Specs_Master.xlsx.

The specs master carries a free-text "Benchmarks" column ("0.15% CTR, 40% VCR,
65% Engagement, 70% Viewability"), which can't be sorted, averaged or graded
against. This script normalises those strings into numeric per-metric targets and
emits a four-sheet workbook:

  Benchmarks           one row per sellable placement: targets, canonical format,
                       measurement caveats, and a review flag where the source
                       value needs a human decision.
  Performance Tracker  paste GAM delivery, pick a placement, get PASS/WATCH/MISS.
  Thresholds           the banding levers + target rollups by format.
  Definitions          metric definitions, measurement caveats, open questions.

ROWS below is the transcription of the specs master, one tuple per placement.
It is the thing to edit when a target changes — re-run the script afterwards.
Targets flagged INTERPRETED are judgement calls, not literal reads; the verbatim
source string is preserved in the workbook's "Source" column either way.

Styling (Arial, band headers, category rows, the F2FBEE row fill) is lifted from
the specs master so the two files read as a set.

Usage:
    pip install openpyxl
    python3 scripts/build_ad_benchmarks.py [--out PATH]

The written file has no cached formula values (openpyxl doesn't compute). Excel
recalculates on open; to bake values in headlessly, run it through LibreOffice.
"""

import argparse
import os

# Column F of the specs master, transcribed. V marks "no target set in the source".
# Tuple: (section, product, size, website, location, canonical_format,
#         ctr, viewability, vcr, engagement,
#         viewability_measurable, measurement_note, review_flag, source_verbatim)
V = None

ROWS = [
 ("__CAT__","STANDARD BANNERS",),
 ("STANDARD BANNERS","Medium Rectangle","300x250","newsweek.com","Article Pages","Display",
  0.001,0.70,V,V,"Yes","Standard GPT slot; measures organically in GAM Active View.","","0.10% CTR / 70% Viewability"),
 ("STANDARD BANNERS","Half Page","300x600","newsweek.com","Article Pages","Display",
  0.001,0.70,V,V,"Yes","Standard GPT slot; measures organically in GAM Active View.","","0.10% CTR / 70% Viewability"),
 ("STANDARD BANNERS","Leaderboard","728x90","newsweek.com","Article Pages","Display",
  0.001,0.70,V,V,"Yes","Standard GPT slot; measures organically in GAM Active View.","","0.10% CTR / 70% Viewability"),
 ("STANDARD BANNERS","Billboard ad","970x250","newsweek.com","Article Pages","Display",
  0.001,0.70,V,V,"Yes","Standard GPT slot; measures organically in GAM Active View.","","0.10% CTR / 70% Viewability"),
 ("STANDARD BANNERS","Smartphone Banner","300x50 / 320x50","newsweek.com","Article Pages","Display",
  0.001,0.70,V,V,"Yes","Standard GPT slot; measures organically in GAM Active View.","","0.10% CTR / 70% Viewability"),
 ("STANDARD BANNERS","Insights Premium Spotlight","1200x675 / Logo","newsweek.com","Homepage","Display",
  0.005,V,V,V,"Yes","Native-style unit; renders in-frame.",
  "No viewability target set - every other standard banner carries 70%. Confirm intended.","0.50% CTR"),
 ("__CAT__","VIDEO PRE-ROLL",),
 ("VIDEO PRE-ROLL","Pre Roll","16:9","newsweek.com","Article Pages","Video",
  0.006,0.70,0.70,V,"Yes","In-player video; VCR from GAM video metrics.",
  "0.60% CTR is high for pre-roll - confirm this is the sold target, not a stretch goal.","0.60% CTR / 70% Viewability / 70% VCR"),
 ("__CAT__","CENTERSTAGE",),
 ("CENTERSTAGE","Desktop Ribbon","3840x350","newsweek.com","Homepage","Centerstage",
  0.01,0.70,V,V,"Yes","In-frame takeover render; measured 58-67% viewable in production.","","1% CTR / 70% Viewability"),
 ("CENTERSTAGE","Desktop Large Banner","3200x700","newsweek.com","Homepage","Centerstage",
  0.001,0.70,V,V,"Yes","In-frame takeover render.",
  "0.10% CTR vs 1% on Ribbon / Mobile Sticky in the same family - 10x gap, confirm.","0.10% CTR / 70% Viewability"),
 ("CENTERSTAGE","Mobile Sticky","900x350","newsweek.com","Homepage","Centerstage",
  0.01,0.70,V,V,"Yes","In-frame takeover render.","","1% CTR / 70% Viewability"),
 ("CENTERSTAGE","Mobile Large Banner","1200x900","newsweek.com","Homepage","Centerstage",
  0.001,0.70,V,V,"Yes","In-frame takeover render.",
  "0.10% CTR vs 1% on Ribbon / Mobile Sticky in the same family - 10x gap, confirm.","0.10% CTR / 70% Viewability"),
 ("CENTERSTAGE","Video Takeover Desktop","16:9","newsweek.com","Homepage","Centerstage",
  0.001,0.70,V,V,"Yes","Video takeover; start/complete tracked.",
  "Video unit with no VCR target - add one to grade completion.","0.10% CTR / 70% Viewability"),
 ("CENTERSTAGE","Video Takeover Mobile","9:16","newsweek.com","Homepage","Centerstage",
  0.001,0.70,V,V,"Yes","Video takeover; start/complete tracked.",
  "Video unit with no VCR target - add one to grade completion.","0.10% CTR / 70% Viewability"),
 ("__CAT__","FITO DISPLAY (FIRST IMPRESSION TAKE OVER)",),
 ("FITO DISPLAY","Desktop Top Banner","970x250","newsweek.com","Article Pages","FITO",
  0.005,V,V,V,"Yes - verify","Served via carrier-slot placement injection; confirm Active View reads real geometry on the live creative.",
  "No viewability target set - sibling FITO rows all carry 70%.","0.50% CTR (stored as a raw number in the specs master)"),
 ("FITO DISPLAY","Mobile Sticky","1300x350 or 320x50","newsweek.com","Article Pages","FITO",
  0.001,0.70,V,V,"Yes","Standard sticky render.","","0.10% CTR / 70% Viewability"),
 ("FITO DISPLAY","Fluid 250 (in-article, optional)","2560x250","newsweek.com","Article Pages","Display",
  V,0.70,V,V,"Yes","Fluid in-article render.",
  "No CTR target set.","70% Viewability"),
 ("FITO DISPLAY","In-Article (Desktop)","3200x500, 1600x250 or 970x250","newsweek.com","Article Pages","Display",
  0.001,0.70,V,V,"Yes","Standard in-article slot.","","0.10% CTR / 70% Viewability"),
 ("FITO DISPLAY","Mobile Top Banner","1200x900 or 300x250","newsweek.com","Article Pages","FITO",
  0.005,V,V,V,"Yes - verify","Served via carrier-slot placement injection; confirm Active View reads real geometry on the live creative.",
  "No viewability target set - sibling FITO rows all carry 70%.","0.50% CTR (stored as a raw number in the specs master)"),
 ("FITO DISPLAY","Medium Rectangle","300x250","newsweek.com","Article Pages","Display",
  0.001,0.70,V,V,"Yes","Standard GPT slot.","","0.10% CTR / 70% Viewability"),
 ("FITO DISPLAY","Leaderboard","728x90","newsweek.com","Article Pages","Display",
  0.001,0.70,V,V,"Yes","Standard GPT slot.","","0.10% CTR / 70% Viewability"),
 ("__CAT__","FITO VIDEO (FIRST IMPRESSION TAKE OVER)",),
 ("FITO VIDEO","Desktop Top Banner","2310x500, 3200x500 or 970x250","newsweek.com","Article Pages","FITO",
  0.006,0.70,V,V,"Yes - verify","In-banner video via injected carrier slot; confirm Active View geometry.",
  "In-banner video with no VCR target - add one to grade completion.","0.60% CTR / 70% Viewability"),
 ("FITO VIDEO","Mobile Leaderboard (Sticky)","1300x350 or 320x50","newsweek.com","Article Pages","FITO",
  0.006,0.70,V,V,"Yes","Standard sticky render.","","0.60% CTR / 70% Viewability"),
 ("FITO VIDEO","Mobile Top Banner","1200x900 or 300x250","newsweek.com","Article Pages","FITO",
  0.001,0.70,V,V,"Yes - verify","Served via carrier-slot placement injection; confirm Active View geometry.",
  "0.10% CTR here vs 0.60% on the other FITO Video units - confirm.","0.10% CTR / 70% Viewability"),
 ("FITO VIDEO","Fluid 250 (in-article, optional)","2560x250","newsweek.com","Article Pages","Display",
  V,0.70,V,V,"Yes","Fluid in-article render.","No CTR target set.","70% Viewability"),
 ("FITO VIDEO","In-Article (Desktop)","3200x500, 1600x250 or 970x250","newsweek.com","Article Pages","Display",
  0.001,0.70,V,V,"Yes","Standard in-article slot.","","0.10% CTR / 70% Viewability"),
 ("FITO VIDEO","Medium Rectangle (mobile)","300x250","newsweek.com","Article Pages","Display",
  0.001,0.70,V,V,"Yes","Standard GPT slot.","","0.10% CTR / 70% Viewability"),
 ("FITO VIDEO","Desktop Leaderboard (Sticky)","728x90","newsweek.com","Article Pages","Display",
  0.001,0.70,V,V,"Yes","Standard GPT slot.","","0.10% CTR / 70% Viewability"),
 ("FITO VIDEO","Video Takeover","Horizontal 16:9","newsweek.com","Homepage","Centerstage",
  0.001,0.70,V,V,"Yes","Video takeover; start/complete tracked.",
  "Video unit with no VCR target - add one to grade completion.","0.10% CTR / 70% Viewability"),
 ("__CAT__","HIGH IMPACT",),
 ("HIGH IMPACT","Fluid 250","2560x600","newsweek.com","Article Pages","Display",
  0.0015,0.70,0.40,0.65,"Yes","Engagement rate is vendor-reported, not a GAM metric - pull it from the high-impact partner's report.",
  "","0.15% CTR, 40% VCR, 65% Engagement, 70% Viewability"),
 ("HIGH IMPACT","Fluid 400","2560x600","newsweek.com","Article Pages","Display",
  0.002,0.50,0.45,0.60,"Yes","Engagement rate is vendor-reported, not a GAM metric.",
  "","0.20% CTR, 45% VCR, 60% Engagement, 50% Viewability"),
 ("HIGH IMPACT","Fluid 600","2560x600","newsweek.com","Article Pages","Display",
  0.0025,0.50,0.45,0.60,"Yes","Engagement rate is vendor-reported, not a GAM metric.",
  "","0.25% CTR, 45% VCR, 60% Engagement, 50% Viewability"),
 ("HIGH IMPACT","Responsive","2560x853","newsweek.com","Article Pages","Display",
  0.0015,0.65,0.55,0.35,"Yes","Engagement rate is vendor-reported, not a GAM metric.",
  "","0.15% CTR, 55% VCR, 35% Engagement, 65% Viewability"),
 ("HIGH IMPACT","Page Scroller","1920x1080","newsweek.com","Article Pages","Interscroller",
  0.0035,0.40,0.65,0.70,"Yes - verify","Scroller units can render outside the GPT iframe; confirm Active View is reading real geometry.",
  "","0.35% CTR, 65% VCR, 70% Engagement, 40% Viewability"),
 ("HIGH IMPACT","Display Interscroller","700x1600","newsweek.com","Article Pages (Mobile Only)","Interscroller",
  0.004,V,V,V,"Yes - iframe mirror required","Breakout render: Active View read ~0% until the iframe-mirror fix was applied (then 34-57%). Confirm the mirror is live on the creative before grading viewability. Never sell a vCPM goal on this unit.",
  "No viewability target set - decide one now that the mirror makes the metric real.","0.40% CTR"),
 ("HIGH IMPACT","Video Interscroller","Vertical 9:16","newsweek.com","Article Pages (Mobile Only)","Interscroller",
  0.01,0.70,V,V,"Yes - iframe mirror required","Breakout render: same iframe-mirror caveat as Display Interscroller. 60-second creative carries a lower 40% viewability target.",
  "INTERPRETED: source reads '0.70% viewability / 0.40% (60 seconds)'. Entered as 70% (and 40% at :60) - a 0.70% viewability target is not plausible. Confirm.",
  "1.00% CTR / 0.70% viewability (30 seconds) / 0.40% (60 seconds)"),
 ("__CAT__","SOCIAL",),
 ("SOCIAL","Facebook Post","600x600","www.facebook.com","Facebook Placements (feed, stories)","Social",
  0.009,V,V,V,"Platform-reported","Meta-reported metrics only; no GAM Active View, no DV.",
  "0.90% CTR vs 0.10% on Instagram - 9x gap between two feed placements, confirm.","0.90% CTR"),
 ("SOCIAL","Instagram Post","600x600","www.instagram.com","Instagram Placements (feed, stories)","Social",
  0.001,V,V,V,"Platform-reported","Meta-reported metrics only.","","0.10% CTR"),
 ("SOCIAL","Instagram and Facebook Video","9:16","www.facebook.com / www.instagram.com","Instagram & Facebook Placements (feed, stories)","Social",
  V,V,V,V,"Platform-reported","Meta-reported metrics only.","NO BENCHMARK SET - needs a CTR and a VCR target.",""),
 ("SOCIAL","LinkedIn Display","1200x627","www.linkedin.com","Feed","Social",
  0.008,V,V,V,"Platform-reported","LinkedIn-reported metrics only.","","0.80% CTR (stored as a raw number in the specs master)"),
 ("SOCIAL","LinkedIn Video","1:1 (feed) / 9:16 (mobile)","www.linkedin.com","Feed","Social",
  0.01,V,V,V,"Platform-reported","LinkedIn-reported metrics only.",
  "Video unit with no VCR target.","1.00% CTR (stored as a raw number in the specs master)"),
 ("SOCIAL","TikTok Display","1200x628","www.tiktok.com","Feed","Social",
  V,V,V,V,"Platform-reported","TikTok-reported metrics only.","NO BENCHMARK SET - needs a CTR target.",""),
 ("SOCIAL","TikTok Video","Vertical 9:16","www.tiktok.com","Feed","Social",
  V,V,V,V,"Platform-reported","TikTok-reported metrics only.","NO BENCHMARK SET - needs a CTR and a VCR target.",""),
 ("__CAT__","NEWSLETTER",),
 ("NEWSLETTER","Medium Rectangle","300x250","newsweek.com","Email Newsletter Banner","Newsletter",
  0.000006,V,V,V,"N/A","Email render - no viewability measurement is possible. CTR is ESP-reported.",
  "0.0006% CTR is ~370x below the Long Banner target in the same newsletter. Almost certainly a decimal error - confirm the intended value (0.06%? 0.6%?).","0.0006% CTR"),
 ("NEWSLETTER","Long Banner","970x550","newsweek.com","Email Newsletter Banner","Newsletter",
  0.0022,V,V,V,"N/A","Email render - no viewability measurement. CTR is ESP-reported.","","0.22% CTR"),
 ("NEWSLETTER","Logo","62px height, 200-400px width","newsweek.com","Email Newsletter Logo","Newsletter",
  0.004,V,V,V,"N/A","Email render - no viewability measurement. CTR is ESP-reported.","","0.40% CTR"),
 ("NEWSLETTER","Native","600x314","newsweek.com","Email Newsletter Native","Newsletter",
  0.006,V,V,V,"N/A","Email render - no viewability measurement. CTR is ESP-reported.","","0.60% CTR"),
 ("__CAT__","PODCAST",),
 ("PODCAST","Podcast Audio Ads","up to 1-minute host-read","newsweek.com","Article Pages + Podcast Section","Podcast",
  V,V,V,V,"N/A","Audio - no CTR or viewability. Grade on delivered impressions and completion rate from the podcast host.",
  "Specs master reads 'N/A'. Consider setting a listen-through-rate target instead.","N/A"),
 ("__CAT__","APPLE NEWS",),
 ("APPLE NEWS","Standard","300x250","apple.news","Article Pages","Apple News",
  0.002,1.00,V,V,"Apple-reported","100% viewability is Apple's guaranteed-view delivery model, not a GAM Active View measurement. Do not compare it to on-site viewability.",
  "Source reads '0.20 CTR' with no % sign - entered as 0.20%.","0.20 CTR / 100% Viewability"),
 ("APPLE NEWS","Double","1242x332 and 1536x264","apple.news","Article Pages","Apple News",
  0.003,1.00,V,V,"Apple-reported","Apple guaranteed-view model, not GAM Active View.",
  "Source reads '0.30 CTR' with no % sign - entered as 0.30%.","0.30 CTR / 100% Viewability"),
 ("APPLE NEWS","Large","1242x699 and 1536x864","apple.news","Article Pages","Apple News",
  0.003,1.00,V,V,"Apple-reported","Apple guaranteed-view model, not GAM Active View.",
  "Source reads '0.30 CTR' with no % sign - entered as 0.30%.","0.30 CTR / 100% Viewability"),
 ("APPLE NEWS","MREC","900x750","apple.news","Homepage + Article Pages","Apple News",
  0.003,1.00,V,V,"Apple-reported","Apple guaranteed-view model, not GAM Active View.",
  "Source reads '0.30 CTR' with no % sign - entered as 0.30%.","0.30 CTR / 100% Viewability"),
 ("APPLE NEWS","Interstitial","1242x2208 and 1536x2048","apple.news","Article Pages","Apple News",
  0.003,1.00,V,V,"Apple-reported","Apple guaranteed-view model, not GAM Active View.",
  "Source reads '0.30 CTR' with no % sign - entered as 0.30%.","0.30 CTR / 100% Viewability"),
 ("APPLE NEWS","PreRoll and OutStream","1080x1920, 1080x1080 or 1920x1080","apple.news","Homepage + Article Pages","Apple News",
  0.003,1.00,V,V,"Apple-reported","Apple guaranteed-view model, not GAM Active View.",
  "Video unit with no VCR target. Source reads '0.30 CTR' with no % sign - entered as 0.30%.","0.30 CTR / 100% Viewability"),
 ("APPLE NEWS","Sponsored Content","1200x627","apple.news","Homepage + Article Pages","Apple News",
  0.01,1.00,V,V,"Apple-reported","Apple guaranteed-view model, not GAM Active View.","","1% CTR / 100% Viewability"),
 ("__CAT__","INTERSTITIAL",),
 ("INTERSTITIAL","Mobile","900x1600","newsweek.com","Article Pages","Interstitial",
  0.008,1.00,V,V,"Yes - verify","100% viewability assumes a full-screen guaranteed view; confirm what GAM Active View actually reports before grading against it.",
  "Source reads '0.80 CTR' with no % sign - entered as 0.80%.","0.80 CTR / 100% Viewability"),
 ("INTERSTITIAL","Desktop","1600x900","newsweek.com","Article Pages","Interstitial",
  0.008,1.00,V,V,"Yes - verify","100% viewability assumes a full-screen guaranteed view; confirm what GAM Active View actually reports.",
  "Source reads '0.80 CTR' with no % sign - entered as 0.80%.","0.80 CTR / 100% Viewability"),
]

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

# ---- palette lifted from Newsweek_Digital_Ads_Specs_Master.xlsx ----
C_PLACEMENT = "69645A"; C_TARGET = "158113"; C_MEASURE = "054B4F"; C_GOV = "0D838A"
C_CATEGORY  = "A6A08E"; C_ROW = "F2FBEE"; C_INPUT = "FFF7CC"; C_CALC = "EFEFEF"
INK = "1F1E19"; MUTED = "69645A"
F = "Arial"
thin = Side(style="thin", color="D6D2C4")
BORD = Border(bottom=thin)

def band(ws, row, c1, c2, text, color):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row, c1, text)
    c.font = Font(F, 10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color); c.alignment = Alignment("center","center")
    for i in range(c1, c2+1):
        ws.cell(row, i).fill = PatternFill("solid", fgColor=color)

def hdr(ws, row, col, text, color):
    c = ws.cell(row, col, text)
    c.font = Font(F, 9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment("center","center", wrap_text=True); return c

def title(ws, row, col, text, sub=None, width=17):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+width-1)
    c = ws.cell(row, col, text); c.font = Font(F, 14, bold=True, color=INK)
    c.alignment = Alignment("left","center")
    if sub:
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+width-1)
        s = ws.cell(row+1, col, sub); s.font = Font(F, 9, italic=True, color=MUTED)
        s.alignment = Alignment("left","center")

wb = Workbook()

# =====================================================================
# SHEET 1 — Benchmarks
# =====================================================================
ws = wb.active; ws.title = "Benchmarks"
ws.sheet_view.showGridLines = False
WID = {"A":3,"B":24,"C":22,"D":20,"E":24,"F":20,"G":14,"H":42,"I":11,"J":12,"K":10,
       "L":12,"M":18,"N":56,"O":12,"P":58,"Q":34,"R":13}
for k,v in WID.items(): ws.column_dimensions[k].width = v

title(ws, 1, 2, "Newsweek Digital Ads — Performance Benchmarks",
      "Companion to Newsweek_Digital_Ads_Specs_Master.xlsx. One row per sellable placement; targets normalised out of the specs master's free-text Benchmarks column into numeric, comparable fields.")
ws.row_dimensions[1].height = 24; ws.row_dimensions[2].height = 26

BANDS = [(2,8,"PLACEMENT",C_PLACEMENT),(9,12,"TARGET",C_TARGET),
         (13,14,"MEASUREMENT",C_MEASURE),(15,18,"GOVERNANCE",C_GOV)]
for c1,c2,t,col in BANDS: band(ws, 3, c1, c2, t, col)
ws.row_dimensions[3].height = 20

HEADERS = [(2,"Ad Product",C_PLACEMENT),(3,"Ad Size",C_PLACEMENT),(4,"Website(s)",C_PLACEMENT),
    (5,"Location",C_PLACEMENT),(6,"Section",C_PLACEMENT),(7,"Canonical Format",C_PLACEMENT),
    (8,"Benchmark Key",C_PLACEMENT),
    (9,"CTR target",C_TARGET),(10,"Viewability target",C_TARGET),(11,"VCR target",C_TARGET),
    (12,"Engagement target",C_TARGET),
    (13,"Viewability measurable?",C_MEASURE),(14,"Measurement & delivery notes",C_MEASURE),
    (15,"Targets set?",C_GOV),(16,"Review flag",C_GOV),(17,"Source (verbatim, specs master)",C_GOV),
    (18,"Last reviewed",C_GOV)]
for col,txt,colr in HEADERS: hdr(ws, 4, col, txt, colr)
ws.row_dimensions[4].height = 40
ws.freeze_panes = "C5"

PCT_CTR = '0.00##%'; PCT = '0.0%'
r = 5
first_data = None
for item in ROWS:
    if item[0] == "__CAT__":
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=18)
        c = ws.cell(r, 2, item[1]); c.font = Font(F, 9, bold=True, color="000000")
        c.alignment = Alignment("center","center")
        for i in range(2,19):
            ws.cell(r,i).fill = PatternFill("solid", fgColor=C_CATEGORY)
            ws.cell(r,i).border = BORD
        ws.row_dimensions[r].height = 20; r += 1; continue
    sec,prod,size,site,loc,fmt,ctr,vw,vcr,eng,meas,mnote,flag,src = item
    if first_data is None: first_data = r
    vals = [(2,prod),(3,size),(4,site),(5,loc),(6,sec),(7,fmt),
            (8,f"=IF($B{r}=\"\",\"\",$F{r}&\" · \"&$B{r}&\" · \"&$C{r})"),
            (9,ctr),(10,vw),(11,vcr),(12,eng),(13,meas),(14,mnote),
            (15,f'=IF(COUNT($I{r}:$L{r})=0,"⚠ none",IF(COUNT($I{r}:$L{r})=1,"partial","yes"))'),
            (16,flag),(17,src),(18,date(2026,9,2))]
    for col,v in vals:
        c = ws.cell(r, col, v)
        c.font = Font(F, 9, bold=(col==2), color=INK)
        c.fill = PatternFill("solid", fgColor=C_ROW)
        c.border = BORD
        c.alignment = Alignment("center" if col not in (14,16) else "left","center", wrap_text=True)
    ws.cell(r,9).number_format = PCT_CTR
    for col in (10,11,12): ws.cell(r,col).number_format = PCT
    ws.cell(r,18).number_format = 'yyyy-mm-dd'
    if flag:
        ws.cell(r,16).font = Font(F, 9, color="C41608", bold=True)
    ws.row_dimensions[r].height = 34
    r += 1
last_data = r-1

n = r
ws.cell(n+1, 2, "Legend").font = Font(F, 10, bold=True, color=INK)
notes = [
 "Percentages are stored as fractions (0.001 = 0.10%). Blank = no target set in the specs master; the 'Targets set?' column flags those.",
 "'Benchmark Key' is a formula (Section · Ad Product · Ad Size) and is the unique lookup key used by the Performance Tracker sheet. Don't overwrite it.",
 "'Review flag' lists what needs a human decision — unit-notation ambiguities, missing targets, and internal inconsistencies found in the source file.",
 "Where a review flag says INTERPRETED, the target in this sheet is a judgement call, not a literal read of the source. Column Q holds the verbatim original.",
 "Targets are the SOLD promise. Banding (how far under target counts as WATCH vs MISS) lives on the Thresholds sheet and is applied on the Performance Tracker.",
]
for i,t in enumerate(notes):
    c = ws.cell(n+2+i, 2, "• " + t); c.font = Font(F, 9, color=MUTED)
    c.alignment = Alignment("left","center")
    ws.merge_cells(start_row=n+2+i, start_column=2, end_row=n+2+i, end_column=18)

DATA_FIRST, DATA_LAST = first_data, last_data

# =====================================================================
# SHEET 2 — Performance Tracker
# =====================================================================
tr = wb.create_sheet("Performance Tracker")
tr.sheet_view.showGridLines = False
TW = {"A":3,"B":34,"C":42,"D":12,"E":12,"F":13,"G":10,"H":15,"I":14,"J":11,"K":13,
      "L":11,"M":12,"N":11,"O":11,"P":12,"Q":11,"R":11,"S":11,"T":11,"U":11,"V":11,"W":11,"X":13}
for k,v in TW.items(): tr.column_dimensions[k].width = v
title(tr, 1, 2, "Performance Tracker — actual vs benchmark",
      "Fill the amber cells. Everything else is a formula: rates, the benchmark pulled from the Benchmarks sheet, the gap, and the verdict.", width=23)
tr.row_dimensions[1].height = 24; tr.row_dimensions[2].height = 26

TB = [(2,5,"LINE ITEM",C_PLACEMENT),(6,11,"DELIVERY (paste from GAM)",C_INPUT),
      (12,14,"ACTUAL",C_MEASURE),(15,17,"TARGET (auto)",C_TARGET),
      (18,20,"GAP vs target (pp)",C_GOV),(21,24,"VERDICT",C_GOV)]
for c1,c2,t,col in TB:
    band(tr, 3, c1, c2, t, col)
    if col == C_INPUT:
        for i in range(c1,c2+1): tr.cell(3,i).font = Font(F,10,bold=True,color=INK)
        tr.cell(3,c1).font = Font(F,10,bold=True,color=INK)
tr.row_dimensions[3].height = 20

THDR = [(2,"Line item / campaign",C_PLACEMENT),(3,"Benchmark Key (pick from list)",C_PLACEMENT),
 (4,"Flight start",C_PLACEMENT),(5,"Flight end",C_PLACEMENT),
 (6,"Impressions",C_INPUT),(7,"Clicks",C_INPUT),(8,"Measurable impr.",C_INPUT),
 (9,"Viewable impr.",C_INPUT),(10,"Video starts",C_INPUT),(11,"Video completes",C_INPUT),
 (12,"CTR",C_MEASURE),(13,"Viewability",C_MEASURE),(14,"VCR",C_MEASURE),
 (15,"CTR",C_TARGET),(16,"Viewability",C_TARGET),(17,"VCR",C_TARGET),
 (18,"CTR",C_GOV),(19,"Viewability",C_GOV),(20,"VCR",C_GOV),
 (21,"CTR",C_GOV),(22,"Viewability",C_GOV),(23,"VCR",C_GOV),(24,"Overall",C_GOV)]
for col,txt,colr in THDR:
    c = hdr(tr, 4, col, txt, colr)
    if colr == C_INPUT: c.font = Font(F, 9, bold=True, color=INK)
tr.row_dimensions[4].height = 34
tr.freeze_panes = "F5"

BM = "Benchmarks"
EX = 5; LASTROW = 5 + 20
for row in range(EX, LASTROW+1):
    is_ex = (row == EX)
    ex = ["EXAMPLE — Infiniti Newsmakers (delete this row)",
          "STANDARD BANNERS · Medium Rectangle · 300x250",
          date(2026,8,1), date(2026,8,31), 1250000, 1500, 1180000, 796500, None, None] if is_ex else [None]*10
    for i,col in enumerate(range(2,12)):
        c = tr.cell(row, col, ex[i])
        c.font = Font(F, 9, color="0000FF" if col>=6 else INK, italic=is_ex)
        c.fill = PatternFill("solid", fgColor=C_INPUT if col>=4 or col==3 else "FFFFFF")
        c.border = BORD; c.alignment = Alignment("left" if col in (2,3) else "center","center", wrap_text=(col in (2,3)))
        if col in (4,5): c.number_format = 'yyyy-mm-dd'
        if col >= 6: c.number_format = '#,##0'
    tr.cell(row,2).fill = PatternFill("solid", fgColor=C_INPUT)
    # A target cell that is blank in Benchmarks must come back as "" here, NOT 0 --
    # a 0% target would silently grade every line PASS on that metric.
    def tgt(colletter):
        idx = (f'INDEX({BM}!${colletter}${DATA_FIRST}:${colletter}${DATA_LAST},'
               f'MATCH($C{row},{BM}!$H${DATA_FIRST}:$H${DATA_LAST},0))')
        return f'=IF($C{row}="","",IFERROR(IF({idx}="","",{idx}),""))'
    calc = [
      (12, f'=IFERROR($G{row}/$F{row},"")'),
      (13, f'=IFERROR($I{row}/$H{row},"")'),
      (14, f'=IFERROR($K{row}/$J{row},"")'),
      (15, tgt("I")),
      (16, tgt("J")),
      (17, tgt("K")),
      (18, f'=IF(OR($L{row}="",$O{row}=""),"",($L{row}-$O{row})*100)'),
      (19, f'=IF(OR($M{row}="",$P{row}=""),"",($M{row}-$P{row})*100)'),
      (20, f'=IF(OR($N{row}="",$Q{row}=""),"",($N{row}-$Q{row})*100)'),
      (21, f'=IF(OR($L{row}="",$O{row}=""),"",IF($F{row}<Thresholds!$D$7,"THIN",IF($L{row}>=$O{row},"PASS",IF($L{row}>=$O{row}*Thresholds!$D$6,"WATCH","MISS"))))'),
      (22, f'=IF(OR($M{row}="",$P{row}=""),"",IF($F{row}<Thresholds!$D$7,"THIN",IF($M{row}>=$P{row},"PASS",IF($M{row}>=$P{row}*Thresholds!$D$6,"WATCH","MISS"))))'),
      (23, f'=IF(OR($N{row}="",$Q{row}=""),"",IF($F{row}<Thresholds!$D$7,"THIN",IF($N{row}>=$Q{row},"PASS",IF($N{row}>=$Q{row}*Thresholds!$D$6,"WATCH","MISS"))))'),
      (24, f'=IF(COUNTIF($U{row}:$W{row},"MISS")>0,"MISS",IF(COUNTIF($U{row}:$W{row},"WATCH")>0,"WATCH",IF(COUNTIF($U{row}:$W{row},"PASS")>0,"PASS",IF(COUNTIF($U{row}:$W{row},"THIN")>0,"THIN",""))))'),
    ]
    for col,f in calc:
        c = tr.cell(row, col, f)
        c.font = Font(F, 9, color=INK, bold=(col==24), italic=is_ex)
        c.fill = PatternFill("solid", fgColor=C_CALC)
        c.border = BORD; c.alignment = Alignment("center","center")
        if col == 12: c.number_format = PCT_CTR
        elif col in (13,14): c.number_format = PCT
        elif col == 15: c.number_format = PCT_CTR
        elif col in (16,17): c.number_format = PCT
        elif col in (18,19,20): c.number_format = '+0.00;-0.00;0.00'
    tr.row_dimensions[row].height = 26

dv = DataValidation(type="list", formula1="Thresholds!$B$28:$B$83", allow_blank=True,
                    showDropDown=False, promptTitle="Benchmark Key",
                    prompt="Pick the placement this line item ran as.")
tr.add_data_validation(dv); dv.add(f"C{EX}:C{LASTROW}")

ln = LASTROW + 2
tr.cell(ln, 2, "How to use").font = Font(F, 10, bold=True, color=INK)
hu = [
 "AMBER cells are yours to fill. Grey cells are formulas — don't type over them.",
 "Column C is a dropdown of every Benchmark Key. Pick one and the targets, gaps and verdicts fill in.",
 "Paste Impressions / Clicks / Measurable / Viewable straight from a GAM delivery report; leave the video columns blank on display lines.",
 "Verdicts: PASS at or above target · WATCH within the tolerance band on the Thresholds sheet · MISS below it · THIN when the line hasn't delivered enough impressions for a fair read.",
 "The row above is an example. Delete it before you share the file.",
]
for i,t in enumerate(hu):
    c = tr.cell(ln+1+i, 2, "• " + t); c.font = Font(F, 9, color=MUTED)
    c.alignment = Alignment("left","center")
    tr.merge_cells(start_row=ln+1+i, start_column=2, end_row=ln+1+i, end_column=14)

# =====================================================================
# SHEET 3 — Thresholds
# =====================================================================
th = wb.create_sheet("Thresholds")
th.sheet_view.showGridLines = False
for k,v in {"A":3,"B":46,"C":16,"D":14,"E":14,"F":14,"G":14,"H":58}.items():
    th.column_dimensions[k].width = v
title(th, 1, 2, "Thresholds, banding & rollups",
      "The tolerance levers the Performance Tracker grades against, plus target averages by format and by section.", width=7)
th.row_dimensions[1].height = 24; th.row_dimensions[2].height = 26

band(th, 4, 2, 4, "BANDING LEVERS — edit these", C_TARGET)
for col,t in [(2,"Lever"),(3,"Value"),(4,"")]:
    pass
th.cell(5,2,"Lever").font = Font(F,9,bold=True,color="FFFFFF")
for i,(lbl,val,fmtn,note) in enumerate([
    ("WATCH band — a metric at or above this share of target is WATCH, below it is MISS", 0.90, '0%',
     "0.90 means: 63% viewability against a 70% target = WATCH; 62% = MISS."),
    ("Minimum impressions for a valid read — below this the verdict is THIN", 10000, '#,##0',
     "Stops a 200-impression line from being graded a MISS on noise."),
    ("Benchmark review cadence (months)", 6, '0',
     "How often this sheet's targets should be re-checked against actual delivery."),
]):
    r2 = 6+i
    c = th.cell(r2, 2, lbl); c.font = Font(F, 9, color=INK); c.alignment = Alignment("left","center", wrap_text=True)
    c.fill = PatternFill("solid", fgColor=C_ROW); c.border = BORD
    v = th.cell(r2, 4, val); v.font = Font(F, 10, bold=True, color="0000FF")
    v.fill = PatternFill("solid", fgColor=C_INPUT); v.number_format = fmtn
    v.alignment = Alignment("center","center"); v.border = BORD
    n2 = th.cell(r2, 8, note); n2.font = Font(F, 9, italic=True, color=MUTED)
    n2.alignment = Alignment("left","center", wrap_text=True)
    th.row_dimensions[r2].height = 30
th.cell(5,2,"").value = None

FORMATS = ["Display","Video","Interstitial","Interscroller","FITO","Centerstage","Apple News","Social","Newsletter","Podcast"]
band(th, 10, 2, 6, "TARGET ROLLUP BY CANONICAL FORMAT (auto)", C_MEASURE)
for col,t in [(2,"Canonical Format"),(3,"Placements"),(4,"Avg CTR target"),(5,"Avg viewability target"),(6,"Avg VCR target")]:
    hdr(th, 11, col, t, C_MEASURE)
th.row_dimensions[11].height = 30
for i,f in enumerate(FORMATS):
    r2 = 12+i
    cells = [(2,f),
      (3,f'=COUNTIF({BM}!$G${DATA_FIRST}:$G${DATA_LAST},$B{r2})'),
      (4,f'=IFERROR(AVERAGEIF({BM}!$G${DATA_FIRST}:$G${DATA_LAST},$B{r2},{BM}!$I${DATA_FIRST}:$I${DATA_LAST}),"—")'),
      (5,f'=IFERROR(AVERAGEIF({BM}!$G${DATA_FIRST}:$G${DATA_LAST},$B{r2},{BM}!$J${DATA_FIRST}:$J${DATA_LAST}),"—")'),
      (6,f'=IFERROR(AVERAGEIF({BM}!$G${DATA_FIRST}:$G${DATA_LAST},$B{r2},{BM}!$K${DATA_FIRST}:$K${DATA_LAST}),"—")')]
    for col,v in cells:
        c = th.cell(r2, col, v); c.font = Font(F, 9, bold=(col==2), color=INK)
        c.fill = PatternFill("solid", fgColor=C_ROW); c.border = BORD
        c.alignment = Alignment("left" if col==2 else "center","center")
        if col == 4: c.number_format = PCT_CTR
        elif col in (5,6): c.number_format = PCT
        elif col == 3: c.number_format = '0'
    th.row_dimensions[r2].height = 20
th.cell(23, 2, "Averages ignore blanks, so a format whose placements have no viewability target reads '—' rather than 0%.").font = Font(F, 9, italic=True, color=MUTED)

band(th, 25, 2, 4, "BENCHMARK KEY LIST — validation source for the Performance Tracker dropdown", C_GOV)
th.cell(26, 2, "Maintenance: add a row to Benchmarks → paste its Benchmark Key here too, or it won't appear in the dropdown.").font = Font(F, 9, italic=True, color="C41608")
th.merge_cells(start_row=26, start_column=2, end_row=26, end_column=4)
keys = [f"{r0[0]} · {r0[1]} · {r0[2]}" for r0 in ROWS if r0[0] != "__CAT__"]
for i,k in enumerate(keys):
    c = th.cell(28+i, 2, k); c.font = Font(F, 9, color=MUTED); c.alignment = Alignment("left","center")

# =====================================================================
# SHEET 4 — Definitions
# =====================================================================
de = wb.create_sheet("Definitions")
de.sheet_view.showGridLines = False
for k,v in {"A":3,"B":30,"C":110}.items(): de.column_dimensions[k].width = v
title(de, 1, 2, "Definitions, caveats & change log", "Read this before quoting a number to a client.", width=2)
de.row_dimensions[1].height = 24; de.row_dimensions[2].height = 26

def section(row, label, pairs, color):
    band(de, row, 2, 3, label, color)
    de.row_dimensions[row].height = 20
    r2 = row + 1
    for a,b in pairs:
        ca = de.cell(r2, 2, a); ca.font = Font(F, 9, bold=True, color=INK)
        ca.fill = PatternFill("solid", fgColor=C_ROW); ca.border = BORD
        ca.alignment = Alignment("left","center", wrap_text=True)
        cb = de.cell(r2, 3, b); cb.font = Font(F, 9, color=INK)
        cb.fill = PatternFill("solid", fgColor=C_ROW); cb.border = BORD
        cb.alignment = Alignment("left","center", wrap_text=True)
        de.row_dimensions[r2].height = 32
        r2 += 1
    return r2 + 1

r3 = section(4, "METRIC DEFINITIONS", [
 ("CTR","Clicks ÷ impressions. On social and newsletter this is platform- or ESP-reported and is not directly comparable to on-site CTR."),
 ("Viewability","Viewable impressions ÷ measurable impressions, per the MRC standard: 50% of pixels in view for 1 continuous second (2 seconds for video). Not the same as impressions delivered."),
 ("Measurable rate","Measurable ÷ total impressions. A low measurable rate makes the viewability number unreliable — check it before grading."),
 ("VCR","Video completes ÷ video starts. Grade long-form pre-roll (>:30) against its own line, not the standard video target."),
 ("Engagement rate","Vendor-defined interaction rate on high-impact units. It comes from the high-impact partner's report, not GAM — confirm the definition per vendor before promising it."),
 ("SIVT / GIVT","Invalid traffic per MRC. Impression-weighted: fraud-classified monitored ads ÷ all monitored ads. Industry bands: green <1%, amber 1–3%, red ≥3%."),
 ("Attention Index","DoubleVerify Authentic Attention, indexed to a 100 baseline. Above 100 is better than the benchmark set."),
], C_TARGET)

r3 = section(r3, "MEASUREMENT CAVEATS — these change what a number means", [
 ("Breakout / parent-DOM formats","Any creative that renders outside the GPT slot iframe reads ~0% viewable in Active View, and DV agrees, because both instrument the hidden iframe. The tell is healthy CTR with more clicks than 'viewable' impressions. Never sell a vCPM goal on one of these units."),
 ("Interscroller (Mobkoi)","Fixed in production by the iframe-mirror creative: same LI went from 0.51% to 56.81% viewable, ~100% measurable. Grade viewability on these units only once the mirror is confirmed live."),
 ("Apple News 100% viewability","Apple's guaranteed-view delivery model, reported by Apple. It is not a GAM Active View measurement and should not be averaged with on-site viewability."),
 ("Social placements","Meta / LinkedIn / TikTok report their own impressions and clicks. No Active View, no DV. Treat these targets as a separate book."),
 ("Newsletter","Email renders have no viewability measurement at all. CTR is ESP-reported and counts differently from ad-server CTR."),
 ("DoubleVerify lag","DV's export runs about two days behind. A line that started yesterday will show '—' for Attention and IVT — that's timing, not underperformance."),
 ("New line items","A line in its first day or two has no stable rate. The THIN verdict on the tracker exists so nobody escalates on 200 impressions."),
], C_MEASURE)

r3 = section(r3, "WHAT NEEDS A DECISION — carried from the specs master", [
 ("Video Interscroller viewability","Source reads '0.70% viewability (30 seconds) / 0.40% (60 seconds)'. Entered here as 70% and 40%. Confirm — a 0.70% viewability target isn't plausible."),
 ("Newsletter Medium Rectangle CTR","Source reads 0.0006%, roughly 370x below the Long Banner target in the same newsletter. Almost certainly a decimal error."),
 ("Apple News and Interstitial CTR","Source reads '0.20 CTR' / '0.30 CTR' / '0.80 CTR' with no % sign. Entered as percentages."),
 ("Missing targets","FITO Desktop and Mobile Top Banner have no viewability target while every sibling FITO row carries 70%. Display Interscroller has none either. Fluid 250 has no CTR target."),
 ("No benchmark at all","Instagram/Facebook Video, TikTok Display, TikTok Video and Podcast Audio have no target set."),
 ("Video units with no VCR","Centerstage Video Takeover (desktop, mobile), FITO Video Desktop Top Banner, LinkedIn Video and Apple News PreRoll all carry a CTR and viewability target but no completion target."),
 ("Internal inconsistencies","Centerstage Ribbon and Mobile Sticky target 1% CTR while Large Banner units in the same family target 0.10%. Facebook targets 0.90% against Instagram's 0.10%. Both gaps are worth a sanity check."),
], C_GOV)

r3 = section(r3, "CELL LEGEND", [
 ("Amber fill, blue text","An input. Type here."),
 ("Grey fill","A formula. Don't type over it."),
 ("Light green fill","Reference data from the specs master."),
 ("Red text in 'Review flag'","Needs a human decision before the number is quoted to a client."),
], C_PLACEMENT)

section(r3, "CHANGE LOG", [
 ("2026-09-02","Created from Newsweek_Digital_Ads_Specs_Master.xlsx. Free-text Benchmarks column parsed into numeric CTR / viewability / VCR / engagement targets; measurement caveats and review flags added."),
 ("","")], C_PLACEMENT)

_repo = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_default = os.path.join(_repo, "docs", "Newsweek_Digital_Ads_Benchmarks_Master.xlsx")
_ap = argparse.ArgumentParser(description=__doc__)
_ap.add_argument("--out", default=_default, help="output .xlsx path")
out = _ap.parse_args().out
wb.save(out)
print(f"wrote {out} ({sum(1 for r in ROWS if r[0] != '__CAT__')} placements, "
      f"benchmark rows {DATA_FIRST}-{DATA_LAST})")
