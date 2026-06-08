"""Update GAM custom targeting key 'bmb' with Bombora topic taxonomy values.

Usage:
    # Add new topics from a committed CSV (default — used by GitHub Actions):
    python scripts/update_bmb_kvps.py --csv data/bombora_penguin_20260608.csv

    # Add new topics from the full Excel file:
    python scripts/update_bmb_kvps.py --xlsx /path/to/Bombora_Topic_Taxonomy.xlsx --sheet full

    # Fix display names on already-created values (safe to re-run):
    python scripts/update_bmb_kvps.py --csv data/bombora_penguin_20260608.csv --fix-display-names

    # Dry-run (no writes):
    python scripts/update_bmb_kvps.py --csv data/bombora_penguin_20260608.csv --dry-run

Display name format: "Theme - Category - Topic Name"
  e.g. "Business - Operations - Submission Triage"

The script diffs the taxonomy against existing GAM values and only creates
what's missing, so it's safe to re-run.  --fix-display-names updates the
displayName of any existing value whose format doesn't match.
"""

import argparse
import csv
import json
import logging
import os
import sys
import tempfile
import time

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)

BATCH_SIZE = 500  # GAM recommends ≤500 objects per create/update call


def _display_name(theme: str, category: str, topic_name: str) -> str:
    return f"{theme} - {category} - {topic_name}"


def load_taxonomy_csv(path: str) -> list[tuple[str, str, str, str]]:
    """Return list of (topic_id, topic_name, theme, category) from CSV."""
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            tid   = str(row.get("topic_id",    row.get("Topic_ID",    ""))).strip()
            name  = str(row.get("topic_name",  row.get("Topic_Name",  ""))).strip()
            theme = str(row.get("theme",       row.get("Theme",       ""))).strip()
            cat   = str(row.get("category",    row.get("Category",    ""))).strip()
            if tid and name:
                rows.append((tid, name, theme, cat))
    return rows


def load_taxonomy_xlsx(path: str, sheet: str) -> list[tuple[str, str, str, str]]:
    """Return list of (topic_id, topic_name, theme, category) from Excel."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        sys.exit("openpyxl not installed — pip install openpyxl")

    sheet_map = {"full": "Bombora Full Taxonomy", "penguin": "PenguinTopics"}
    sheet_name = sheet_map.get(sheet, sheet)
    wb = openpyxl.load_workbook(path, read_only=True)
    if sheet_name not in wb.sheetnames:
        sys.exit(f"Sheet '{sheet_name}' not found; available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = []
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = row
            continue
        if len(row) < 4:
            continue
        # Theme, Category, Topic_ID, Topic_Name, ...
        theme = str(row[0]).strip() if row[0] is not None else ""
        cat   = str(row[1]).strip() if row[1] is not None else ""
        tid   = str(row[2]).strip() if row[2] is not None else ""
        name  = str(row[3]).strip() if row[3] is not None else ""
        if tid and name:
            rows.append((tid, name, theme, cat))
    return rows


def get_soap_client():
    sa_json = os.environ.get("GAM_SERVICE_ACCOUNT_JSON")
    if not sa_json:
        sys.exit("GAM_SERVICE_ACCOUNT_JSON env var not set")
    network_id = os.environ.get("GAM_NETWORK_ID")
    if not network_id:
        sys.exit("GAM_NETWORK_ID env var not set")

    try:
        from googleads import ad_manager, oauth2  # type: ignore
    except ImportError:
        sys.exit("googleads not installed — pip install googleads")

    key_data = json.loads(sa_json)
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
        json.dump(key_data, f)
        key_file = f.name

    oauth2_client = oauth2.GoogleServiceAccountClient(
        key_file, "https://www.googleapis.com/auth/dfp"
    )
    client = ad_manager.AdManagerClient(
        oauth2_client, "NewsweekDashboard/1.0", network_code=network_id
    )
    return client, ad_manager


def find_key_id(svc, ad_manager, key_name: str) -> int:
    sb = ad_manager.StatementBuilder(version="v202605")
    sb.Where("name = :name AND status = 'ACTIVE'")
    sb.WithBindVariable("name", key_name)
    sb.Limit(10)
    resp = svc.getCustomTargetingKeysByStatement(sb.ToStatement())
    results = getattr(resp, "results", None) or []
    if not results:
        sys.exit(f"No active custom targeting key named '{key_name}' found.")
    if len(results) > 1:
        log.warning("Multiple keys named '%s' — using first (id=%s)", key_name, results[0].id)
    key_id = int(results[0].id)
    log.info("Found key '%s' → id=%d", key_name, key_id)
    return key_id


def fetch_existing_values(svc, ad_manager, key_id: int) -> dict[str, dict]:
    """Return {value_name: {id, displayName}} for all active values on this key."""
    existing: dict[str, dict] = {}
    sb = ad_manager.StatementBuilder(version="v202605")
    sb.Where("customTargetingKeyId = :kid AND status = 'ACTIVE'")
    sb.WithBindVariable("kid", key_id)
    sb.Limit(500)
    while True:
        resp = svc.getCustomTargetingValuesByStatement(sb.ToStatement())
        results = getattr(resp, "results", None) or []
        if not results:
            break
        for v in results:
            vname = str(getattr(v, "name", ""))
            existing[vname] = {
                "id":          int(getattr(v, "id", 0)),
                "displayName": str(getattr(v, "displayName", "") or ""),
                "matchType":   str(getattr(v, "matchType", "EXACT")),
            }
        sb.offset += sb.limit
        if sb.offset >= getattr(resp, "totalResultSetSize", 0):
            break
    log.info("Existing values in GAM for key id=%d: %d", key_id, len(existing))
    return existing


def create_values_batch(
    svc, key_id: int, batch: list[tuple[str, str, str, str]], dry_run: bool
) -> int:
    """Create a batch of (topic_id, topic_name, theme, category) values."""
    if dry_run:
        log.info("  [dry-run] would create %d values", len(batch))
        return len(batch)

    objects = []
    for tid, name, theme, cat in batch:
        objects.append({
            "customTargetingKeyId": key_id,
            "name":        tid,
            "displayName": _display_name(theme, cat, name)[:255],
            "matchType":   "EXACT",
        })
    result = svc.createCustomTargetingValues(objects)
    return len(result) if result else 0


def update_display_names_batch(
    svc, key_id: int, batch: list[tuple[int, str, str, str, str]], dry_run: bool
) -> int:
    """Update display names for a batch of (gam_id, tid, name, theme, cat) tuples."""
    if dry_run:
        log.info("  [dry-run] would update %d display names", len(batch))
        return len(batch)

    objects = []
    for gam_id, tid, name, theme, cat in batch:
        objects.append({
            "id":                   gam_id,
            "customTargetingKeyId": key_id,
            "name":                 tid,
            "displayName":          _display_name(theme, cat, name)[:255],
            "matchType":            "EXACT",
        })
    result = svc.updateCustomTargetingValues(objects)
    return len(result) if result else 0


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    src = parser.add_mutually_exclusive_group(required=True)
    src.add_argument("--csv",  metavar="PATH", help="CSV file with topic columns")
    src.add_argument("--xlsx", metavar="PATH", help="Excel taxonomy file from Bombora")
    parser.add_argument(
        "--sheet", default="full", choices=["full", "penguin"],
        help="Excel sheet to read (only with --xlsx; default: full)",
    )
    parser.add_argument("--key-name", default="bmb", help="GAM targeting key name (default: bmb)")
    parser.add_argument(
        "--fix-display-names", action="store_true",
        help="Update displayName on already-existing values that don't match the expected format",
    )
    parser.add_argument("--dry-run", action="store_true", help="Report changes without writing")
    args = parser.parse_args()

    if args.csv:
        taxonomy = load_taxonomy_csv(args.csv)
        log.info("Loaded %d topics from CSV %s", len(taxonomy), args.csv)
    else:
        taxonomy = load_taxonomy_xlsx(args.xlsx, args.sheet)
        log.info("Loaded %d topics from Excel %s (sheet=%s)", len(taxonomy), args.xlsx, args.sheet)

    if not taxonomy:
        sys.exit("No topics loaded — check your input file.")

    client, ad_manager = get_soap_client()
    svc = client.GetService("CustomTargetingService", version="v202605")

    key_id   = find_key_id(svc, ad_manager, args.key_name)
    existing = fetch_existing_values(svc, ad_manager, key_id)

    # ── Mode: fix display names on already-existing values ──────────────────
    if args.fix_display_names:
        to_fix: list[tuple[int, str, str, str, str]] = []
        for tid, name, theme, cat in taxonomy:
            if tid not in existing:
                continue
            want = _display_name(theme, cat, name)[:255]
            have = existing[tid]["displayName"]
            if have != want:
                to_fix.append((existing[tid]["id"], tid, name, theme, cat))
                if len(to_fix) <= 5 or args.dry_run:
                    log.info("  update %s: %r → %r", tid, have, want)

        log.info("Values needing display-name fix: %d", len(to_fix))
        if not to_fix:
            log.info("Nothing to fix.")
            return

        total_updated = 0
        for i in range(0, len(to_fix), BATCH_SIZE):
            batch = to_fix[i : i + BATCH_SIZE]
            log.info(
                "Updating batch %d/%d (%d values)…",
                i // BATCH_SIZE + 1, -(-len(to_fix) // BATCH_SIZE), len(batch),
            )
            n = update_display_names_batch(svc, key_id, batch, dry_run=args.dry_run)
            total_updated += n
            if i + BATCH_SIZE < len(to_fix):
                time.sleep(0.5)

        log.info("Done. Updated %d display names.", total_updated)
        return

    # ── Mode: create missing values ─────────────────────────────────────────
    to_add = [(tid, name, theme, cat) for tid, name, theme, cat in taxonomy
              if tid not in existing]
    log.info(
        "Taxonomy: %d  |  Already in GAM: %d  |  To add: %d",
        len(taxonomy), len(existing), len(to_add),
    )

    if not to_add:
        log.info("Nothing to do — GAM is already up to date.")
        return

    if args.dry_run:
        log.info("[dry-run] First 10 values that would be added:")
        for tid, name, theme, cat in to_add[:10]:
            log.info("  %s  →  %s", tid, _display_name(theme, cat, name))
        if len(to_add) > 10:
            log.info("  ... and %d more", len(to_add) - 10)
        log.info("[dry-run] Total that would be added: %d", len(to_add))
        return

    total_created = 0
    for i in range(0, len(to_add), BATCH_SIZE):
        batch = to_add[i : i + BATCH_SIZE]
        log.info(
            "Creating batch %d/%d (%d values)…",
            i // BATCH_SIZE + 1, -(-len(to_add) // BATCH_SIZE), len(batch),
        )
        n = create_values_batch(svc, key_id, batch, dry_run=False)
        total_created += n
        if i + BATCH_SIZE < len(to_add):
            time.sleep(0.5)

    log.info("Done. Created %d new values for key '%s'.", total_created, args.key_name)


if __name__ == "__main__":
    main()
