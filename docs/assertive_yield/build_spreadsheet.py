import os, openpyxl
from collections import OrderedDict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO=os.path.dirname(os.path.abspath(__file__))
SRC=os.path.join(REPO,'source_exports')
FILES={'Desktop':os.path.join(SRC,'DemandManagerWeb_ServerNewsweek_Desktop_ServerPatterns.xlsx'),
       'Mobile':os.path.join(SRC,'DemandManagerWeb_ServerNewsweek_Mobile_ServerPatterns.xlsx')}
CHAN={'puc':'Display','vast':'Video'}
NOTES={'appnexus':'AppNexus / Xandr (Microsoft Monetize)','triplelift':'TripleLift','zeta_global_ssp':'Zeta Global SSP (via franklymedia reseller)','medianet':'Media.net','undertone':'Undertone','rubicon':'Magnite / Rubicon Project','onetag':'OneTag','yahooAds':'Yahoo (yahooAds / yahoossp)','openx':'OpenX','sharethrough':'Sharethrough','minutemedia':'Minute Media','ix':'Index Exchange','imds':'Advangelists / IMDS','insticator':'Insticator','nativo':'Nativo','vidazoo':'Vidazoo','pubmatic':'PubMatic','ogury':'Ogury','sovrn':'Sovrn','kargo':'Kargo','sparteo':'Sparteo','oms':'OMS / Online Media Solutions','openweb':'OpenWeb','smartadserver':'Equativ / Smart AdServer','amx':'AMX RTB','inmobi':'InMobi','seedtag':'Seedtag','mobkoi':'Mobkoi'}

def load(path):
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True); ws=wb['Patterns']
    rows=list(ws.iter_rows(values_only=True)); wb.close()
    ncols=max(len(r) for r in rows)
    grid=[[(rows[r][c] if c<len(rows[r]) else None) for c in range(ncols)] for r in range(len(rows))]
    def cell(r,c):
        v=grid[r][c]; return '' if v is None else str(v).strip()
    adunits=OrderedDict((r,cell(r,0)) for r in range(2,15) if cell(r,0))
    sizes={r:cell(r,2) for r in range(2,15)}
    groups=[];cur=None
    for c in range(3,ncols):
        if cell(0,c): cur={'header':cell(0,c),'cols':[]};groups.append(cur)
        if cur is not None: cur['cols'].append(c)
    parsed=[]
    for g in groups:
        bidder,_,channel=g['header'].partition(':')
        params=OrderedDict()
        for c in g['cols']:
            pn=cell(1,c)
            if not pn: continue
            params[pn]={r:cell(r,c) for r in adunits if cell(r,c)}
        parsed.append({'bidder':bidder,'channel':channel,'params':params})
    return {'adunits':adunits,'sizes':sizes,'groups':parsed}

datas={p:load(f) for p,f in FILES.items()}

# ---- styles ----
HFILL=PatternFill('solid',fgColor='1F1E19'); HFONT=Font(color='FFFFFF',bold=True,size=11)
SUBFILL=PatternFill('solid',fgColor='F0ECE0'); BOLD=Font(bold=True)
THIN=Side(style='thin',color='D8D2C2'); BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
WARN=PatternFill('solid',fgColor='FBE9C9'); CONSTFILL=PatternFill('solid',fgColor='EAF0E2')
CEN=Alignment(horizontal='center',vertical='center'); WRAP=Alignment(vertical='top',wrap_text=True)
TOP=Alignment(vertical='top')

def style_header(ws,row=1,ncol=None):
    ncol=ncol or ws.max_column
    for c in range(1,ncol+1):
        cell=ws.cell(row=row,column=c); cell.fill=HFILL; cell.font=HFONT
        cell.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True); cell.border=BORDER
    ws.row_dimensions[row].height=28
    ws.freeze_panes=ws.cell(row=row+1,column=1)

def widths(ws,ws_widths):
    for col,w in ws_widths.items(): ws.column_dimensions[col].width=w

wb=openpyxl.Workbook()

# ===== Sheet 1: Read me =====
ws=wb.active; ws.title='Read me'
lines=[
 ('Newsweek — SSP / Prebid Server parameters',True,16),
 ('Magnite Demand Manager → Assertive Yield Prebid Server migration',True,12),
 ('',False,11),
 ('Purpose: full server-side (S2S / Prebid Server) demand configuration captured from',False,11),
 ('Magnite Demand Manager so it can be re-entered into Assertive Yield PBS before Magnite',False,11),
 ('is terminated. Source: the Desktop + Mobile "Server Patterns" exports from Demand Manager.',False,11),
 ('',False,11),
 ('Sheets in this workbook:',True,11),
 ('  • Bidder roster — the 28 bidders and where each runs (platform × display/video).',False,11),
 ('  • Account-level IDs — the shared publisher/seller credentials to port first.',False,11),
 ('  • Parameters (tidy) — every parameter value, one row each; filter/pivot freely.',False,11),
 ('  • Desktop / Mobile — per-bidder layout grouped by ad unit (human-readable).',False,11),
 ('',False,11),
 ('Channels: "Display" = puc (banner/native demand); "Video" = vast (VAST demand).',False,11),
 ('Scope column: "constant" = same across all ad units (account-level); "per-ad-unit" = slot-specific.',False,11),
 ('Coverage: a bidder may be wired to only some slots — see the Parameters/Desktop/Mobile sheets.',False,11),
 ('',False,11),
 ('NOT in these exports (inventory separately before cutover): price granularity, auction/bidder',False,11),
 ('timeouts, floors (floor/bidFloor columns are empty in the source), UserID/identity modules, currency.',False,11),
]
for i,(t,b,sz) in enumerate(lines,1):
    c=ws.cell(row=i,column=1,value=t); c.font=Font(bold=b,size=sz)
ws.column_dimensions['A'].width=100

# ===== Sheet 2: Bidder roster =====
ws=wb.create_sheet('Bidder roster')
hdr=['Magnite bidder key','Exchange / notes','Desktop Display','Desktop Video','Mobile Display','Mobile Video','# platforms/channels']
ws.append(hdr)
roster=OrderedDict()
for plat,d in datas.items():
    for g in d['groups']:
        has=any(v for v in g['params'].values())
        e=roster.setdefault(g['bidder'],{'Desktop-puc':False,'Desktop-vast':False,'Mobile-puc':False,'Mobile-vast':False})
        if has: e[f"{plat}-{g['channel']}"]=True
for b,e in roster.items():
    vals=[e['Desktop-puc'],e['Desktop-vast'],e['Mobile-puc'],e['Mobile-vast']]
    ws.append([b,NOTES.get(b,''),*['✓' if v else '' for v in vals],sum(vals)])
style_header(ws)
for r in range(2,ws.max_row+1):
    ws.cell(row=r,column=1).font=BOLD
    for c in range(3,7):
        cell=ws.cell(row=r,column=c); cell.alignment=CEN
        if cell.value=='✓': cell.font=Font(color='3C6B14',bold=True)
    ws.cell(row=r,column=7).alignment=CEN
    for c in range(1,8): ws.cell(row=r,column=c).border=BORDER
widths(ws,{'A':22,'B':42,'C':15,'D':14,'E':14,'F':13,'G':18})
ws.auto_filter.ref=f"A1:G{ws.max_row}"

# ===== build tidy rows =====
def scope_split(d):
    """yield (bidder,channel,param,scope,{adunit:val},present_units,chan_units)"""
    out=[]
    adunits=d['adunits']; sizes=d['sizes']
    for g in d['groups']:
        chan_units=[r for r in adunits if (sizes.get(r)=='video')==(g['channel']=='vast')]
        present=[r for r in adunits if any(r in v for v in g['params'].values())]
        for pn,vals in g['params'].items():
            distinct={vals[r] for r in present if vals.get(r)}
            if not distinct: continue
            const = len(distinct)==1 and (all(vals.get(r) for r in present) or len(present)==1)
            out.append((g['bidder'],g['channel'],pn,'constant' if const else 'per-ad-unit',vals,present,chan_units))
    return out

# ===== Sheet 3: Account-level IDs =====
ws=wb.create_sheet('Account-level IDs')
ws.append(['Platform','Channel','Bidder','Exchange / notes','Parameter','Value (account-level constant)'])
for plat,d in datas.items():
    for bidder,chan,pn,scope,vals,present,chan_units in scope_split(d):
        if scope!='constant': continue
        v=next(iter({vals[r] for r in present if vals.get(r)}))
        ws.append([plat,CHAN[chan],bidder,NOTES.get(bidder,''),pn,v])
style_header(ws)
for r in range(2,ws.max_row+1):
    for c in range(1,7):
        cell=ws.cell(row=r,column=c); cell.border=BORDER; cell.alignment=TOP
    ws.cell(row=r,column=3).font=BOLD
widths(ws,{'A':10,'B':9,'C':20,'D':40,'E':16,'F':46})
ws.auto_filter.ref=f"A1:F{ws.max_row}"

# ===== Sheet 4: Parameters (tidy) =====
ws=wb.create_sheet('Parameters (tidy)')
ws.append(['Platform','Channel','Bidder','Exchange / notes','Ad unit','Size','Parameter','Value','Scope','Full coverage?'])
for plat,d in datas.items():
    adunits=d['adunits']; sizes=d['sizes']
    for bidder,chan,pn,scope,vals,present,chan_units in scope_split(d):
        full = 'yes' if len(present)>=len(chan_units) else 'PARTIAL'
        for r in present:
            if not vals.get(r): continue
            ws.append([plat,CHAN[chan],bidder,NOTES.get(bidder,''),adunits[r],sizes.get(r,''),pn,vals[r],scope,full])
style_header(ws)
for r in range(2,ws.max_row+1):
    for c in range(1,11):
        cell=ws.cell(row=r,column=c); cell.border=BORDER; cell.alignment=TOP
    ws.cell(row=r,column=3).font=BOLD
    if ws.cell(row=r,column=10).value=='PARTIAL':
        ws.cell(row=r,column=10).fill=WARN; ws.cell(row=r,column=10).font=Font(color='8A5A00',bold=True)
    if ws.cell(row=r,column=9).value=='constant':
        ws.cell(row=r,column=9).fill=CONSTFILL
widths(ws,{'A':10,'B':9,'C':18,'D':38,'E':26,'F':17,'G':16,'H':40,'I':12,'J':14})
ws.auto_filter.ref=f"A1:J{ws.max_row}"

# ===== Sheets 5/6: Desktop / Mobile per-bidder readable =====
def build_platform_sheet(plat,d):
    ws=wb.create_sheet(plat)
    adunits=d['adunits']; sizes=d['sizes']
    r=1
    ws.cell(row=r,column=1,value=f'{plat} — ad units').font=Font(bold=True,size=13); r+=2
    ws.cell(row=r,column=1,value='Ad unit').font=BOLD; ws.cell(row=r,column=2,value='Size').font=BOLD
    ws.cell(row=r,column=1).fill=SUBFILL; ws.cell(row=r,column=2).fill=SUBFILL; r+=1
    for ar,nm in adunits.items():
        ws.cell(row=r,column=1,value=nm); ws.cell(row=r,column=2,value=sizes.get(ar,'')); r+=1
    r+=1
    for chan in ['puc','vast']:
        groups=[g for g in d['groups'] if g['channel']==chan and any(v for v in g['params'].values())]
        if not groups: continue
        chan_units=[ar for ar in adunits if (sizes.get(ar)=='video')==(chan=='vast')]
        ws.cell(row=r,column=1,value=f'{CHAN[chan]} bidders ({chan})').font=Font(bold=True,size=12,color='E91D0C'); r+=2
        for g in groups:
            bidder=g['bidder']
            present=[ar for ar in adunits if any(ar in v for v in g['params'].values())]
            # title
            tcell=ws.cell(row=r,column=1,value=f'{bidder}  —  {NOTES.get(bidder,"")}')
            tcell.font=Font(bold=True,size=11)
            for cc in range(1,3): ws.cell(row=r,column=cc).fill=HFILL; ws.cell(row=r,column=cc).font=Font(bold=True,color='FFFFFF')
            tcell.font=Font(bold=True,color='FFFFFF'); r+=1
            if len(present)<len(chan_units):
                wc=ws.cell(row=r,column=1,value='⚠ Partial coverage — only: '+', '.join(adunits[ar] for ar in present))
                wc.font=Font(italic=True,color='8A5A00'); wc.fill=WARN; r+=1
            # constants
            consts=[]; varying=[]
            for pn,vals in g['params'].items():
                distinct={vals[ar] for ar in present if vals.get(ar)}
                if not distinct: continue
                if len(distinct)==1 and (all(vals.get(ar) for ar in present) or len(present)==1):
                    consts.append((pn,next(iter(distinct))))
                else: varying.append((pn,vals))
            if consts:
                ws.cell(row=r,column=1,value='Constant params:').font=Font(italic=True,color='6B6B6B'); r+=1
                for pn,v in consts:
                    ws.cell(row=r,column=1,value=pn).alignment=Alignment(indent=1)
                    vc=ws.cell(row=r,column=2,value=v); vc.fill=CONSTFILL; r+=1
            if varying:
                # header row: Ad unit | param1 | param2 ...
                ws.cell(row=r,column=1,value='Ad unit').font=BOLD; ws.cell(row=r,column=1).fill=SUBFILL
                for i,(pn,_) in enumerate(varying):
                    c=ws.cell(row=r,column=2+i,value=pn); c.font=BOLD; c.fill=SUBFILL
                r+=1
                for ar in present:
                    ws.cell(row=r,column=1,value=adunits[ar])
                    for i,(pn,vals) in enumerate(varying):
                        ws.cell(row=r,column=2+i,value=vals.get(ar,''))
                    r+=1
            r+=1
    # widths
    maxc=ws.max_column
    ws.column_dimensions['A'].width=34
    for ci in range(2,maxc+1): ws.column_dimensions[get_column_letter(ci)].width=24
    ws.freeze_panes='A1'
    return ws

build_platform_sheet('Desktop',datas['Desktop'])
build_platform_sheet('Mobile',datas['Mobile'])

out=os.path.join(REPO,'Newsweek_SSP_Params_Assertive_Yield_Migration.xlsx')
wb.save(out)
print('wrote',out)
# quick stats
import openpyxl as o2
w2=o2.load_workbook(out)
for s in w2.sheetnames: print(' sheet',s,'rows',w2[s].max_row)
